from datetime import datetime, date, timedelta
import io
import json
import base64
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict
from openpyxl.utils import get_column_letter
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User

from django.core.files.base import ContentFile
from django.db.models import Q, Sum, Count
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from .models import Motorista, Veiculo, Mercado, RegistroPonto
from .forms import RegistroPontoForm, MotoristaForm, VeiculoForm, MercadoForm


# =====================
# VIEWS DE AUTENTICAÇÃO
# =====================

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            
            # Redirecionar baseado no tipo de usuário
            if user.is_superuser or user.is_staff:
                return redirect('admin_dashboard')
            else:
                return redirect('motorista_dashboard')
        else:
            messages.error(request, 'Usuário ou senha inválidos!')
    
    return render(request, 'ponto/login.html')

@login_required
def logout_view(request):
    logout(request)
    messages.success(request, 'Logout realizado com sucesso!')
    return redirect('login')

# =====================
# VIEWS DASHBOARD
# =====================

@login_required
def admin_dashboard(request):
    """Dashboard administrativo"""
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Acesso negado!')
        return redirect('motorista_dashboard')
    
    # Estatísticas gerais
    total_motoristas = Motorista.objects.filter(ativo=True).count()
    total_veiculos = Veiculo.objects.filter(ativo=True).count()
    total_mercados = Mercado.objects.filter(ativo=True).count()
    
    # Registros de hoje
    hoje = timezone.now().date()
    registros_hoje = RegistroPonto.objects.filter(data_hora__date=hoje)
    entradas_hoje = registros_hoje.filter(tipo='entrada').count()
    saidas_hoje = registros_hoje.filter(tipo='saida').count()
    
    # Últimos registros
    ultimos_registros = RegistroPonto.objects.select_related(
        'motorista', 'motorista__veiculo'
    ).order_by('-data_hora')[:10]
    
    context = {
        'total_motoristas': total_motoristas,
        'total_veiculos': total_veiculos,
        'total_mercados': total_mercados,
        'entradas_hoje': entradas_hoje,
        'saidas_hoje': saidas_hoje,
        'ultimos_registros': ultimos_registros,
    }
    
    return render(request, 'ponto/admin/admin_dashboard.html', context)

@login_required
def motorista_dashboard(request):
    """Dashboard do motorista"""
    try:
        motorista = request.user.motorista
    except Motorista.DoesNotExist:
        messages.error(request, 'Usuário não está associado a um motorista!')
        return redirect('login')
    
    # Registro de hoje
    hoje = timezone.now().date()
    entrada_hoje = RegistroPonto.objects.filter(
        motorista=motorista,
        tipo='entrada',
        data_hora__date=hoje
    ).first()
    
    saida_hoje = RegistroPonto.objects.filter(
        motorista=motorista,
        tipo='saida',
        data_hora__date=hoje
    ).first()
    
    # Últimos registros
    ultimos_registros = RegistroPonto.objects.filter(
        motorista=motorista
    ).order_by('-data_hora')[:5]
    
    context = {
        'motorista': motorista,
        'entrada_hoje': entrada_hoje,
        'saida_hoje': saida_hoje,
        'ultimos_registros': ultimos_registros,
        'pode_entrada': not entrada_hoje,
        'pode_saida': entrada_hoje and not saida_hoje,
    }
    
    return render(request, 'ponto/motorista_dashboard.html', context)

# =====================
# VIEWS REGISTRO PONTO
# =====================

@login_required
def registrar_ponto(request, tipo):
    """Registra ponto de entrada ou saída"""
    if tipo not in ['entrada', 'saida']:
        messages.error(request, 'Tipo de registro inválido!')
        return redirect('motorista_dashboard')
    
    try:
        motorista = request.user.motorista
    except Motorista.DoesNotExist:
        messages.error(request, 'Usuário não está associado a um motorista!')
        return redirect('login')
    
    # Validações
    hoje = timezone.now().date()
    
    if tipo == 'entrada':
        # Verificar se já fez entrada hoje
        if RegistroPonto.objects.filter(
            motorista=motorista,
            tipo='entrada',
            data_hora__date=hoje
        ).exists():
            messages.error(request, 'Entrada já registrada hoje!')
            return redirect('motorista_dashboard')
    
    elif tipo == 'saida':
        # Verificar se já fez entrada
        entrada_hoje = RegistroPonto.objects.filter(
            motorista=motorista,
            tipo='entrada',
            data_hora__date=hoje
        ).first()
        
        if not entrada_hoje:
            messages.error(request, 'Registre primeiro a entrada!')
            return redirect('motorista_dashboard')
        
        # Verificar se já fez saída
        if RegistroPonto.objects.filter(
            motorista=motorista,
            tipo='saida',
            data_hora__date=hoje
        ).exists():
            messages.error(request, 'Saída já registrada hoje!')
            return redirect('motorista_dashboard')
    
    if request.method == 'POST':
        form = RegistroPontoForm(request.POST, request.FILES)
        if form.is_valid():
            registro = form.save(commit=False)
            registro.motorista = motorista
            registro.tipo = tipo
            
            # Processar fotos com marca d'água
            if 'foto_odometro' in request.FILES:
                registro.foto_odometro = processar_foto_com_marca_dagua(
                    request.FILES['foto_odometro']
                )
            
            if 'foto_combustivel' in request.FILES:
                registro.foto_combustivel = processar_foto_com_marca_dagua(
                    request.FILES['foto_combustivel']
                )
            
            registro.save()
            
            messages.success(
                request, 
                f'{tipo.capitalize()} registrada com sucesso!'
            )
            return redirect('motorista_dashboard')
    else:
        form = RegistroPontoForm()
    
    context = {
        'form': form,
        'tipo': tipo,
        'motorista': motorista,
        'titulo': f'Registrar {tipo.capitalize()}'
    }
    
    return render(request, 'ponto/registrar_ponto.html', context)

def processar_foto_com_marca_dagua(foto):
    """Adiciona marca d'água com data/hora na foto"""
    try:
        # Abrir imagem
        img = Image.open(foto)
        
        # Converter para RGB se necessário
        if img.mode != 'RGB':
            img = img.convert('RGB')
        
        # Criar objeto de desenho
        draw = ImageDraw.Draw(img)
        
        # Data/hora atual
        agora = timezone.now().strftime('%d/%m/%Y %H:%M')
        
        # Configurar fonte (usar fonte padrão se não encontrar)
        try:
            font = ImageFont.truetype("arial.ttf", 36)
        except:
            font = ImageFont.load_default()
        
        # Posição no canto inferior direito
        largura, altura = img.size
        texto_bbox = draw.textbbox((0, 0), agora, font=font)
        texto_largura = texto_bbox[2] - texto_bbox[0]
        texto_altura = texto_bbox[3] - texto_bbox[1]
        
        x = largura - texto_largura - 20
        y = altura - texto_altura - 20
        
        # Adicionar sombra (fundo escuro para legibilidade)
        draw.rectangle(
            [x-10, y-5, x+texto_largura+10, y+texto_altura+5], 
            fill=(0, 0, 0, 180)
        )
        
        # Adicionar texto branco
        draw.text((x, y), agora, fill=(255, 255, 255), font=font)
        
        # Salvar em memória
        output = io.BytesIO()
        img.save(output, format='JPEG', quality=95)
        output.seek(0)
        
        # Retornar como ContentFile
        return ContentFile(
            output.read(),
            name=f"marca_dagua_{foto.name}"
        )
        
    except Exception as e:
        # Em caso de erro, retornar foto original
        print(f"Erro ao processar foto: {e}")
        return foto

# =====================
# VIEWS ADMINISTRATIVAS
# =====================

@login_required
def listar_motoristas(request):
    """Lista todos os motoristas"""
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Acesso negado!')
        return redirect('motorista_dashboard')
    
    motoristas = Motorista.objects.select_related(
        'user', 'veiculo', 'mercado'
    ).order_by('nome_completo')
    
    return render(request, 'ponto/admin/motoristas.html', {
        'motoristas': motoristas
    })

@login_required
def cadastrar_motorista(request):
    """Cadastra novo motorista"""
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Acesso negado!')
        return redirect('motorista_dashboard')
    
    if request.method == 'POST':
        form = MotoristaForm(request.POST)
        if form.is_valid():
            # Criar usuário
            user = User.objects.create_user(
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password'],
                first_name=form.cleaned_data['nome_completo'].split()[0],
                last_name=' '.join(form.cleaned_data['nome_completo'].split()[1:])
            )
            
            # Criar motorista
            motorista = form.save(commit=False)
            motorista.user = user
            motorista.save()
            
            messages.success(request, 'Motorista cadastrado com sucesso!')
            return redirect('listar_motoristas')
    else:
        form = MotoristaForm()
    
    return render(request, 'ponto/admin/cadastrar_motorista.html', {
        'form': form
    })

@login_required
def editar_motorista(request, id):
    """Edita motorista existente"""
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Acesso negado!')
        return redirect('motorista_dashboard')
    
    motorista = get_object_or_404(Motorista, id=id)
    
    if request.method == 'POST':
        form = MotoristaForm(request.POST, instance=motorista)
        if form.is_valid():
            form.save()
            messages.success(request, 'Motorista atualizado com sucesso!')
            return redirect('listar_motoristas')
    else:
        form = MotoristaForm(instance=motorista)
    
    return render(request, 'ponto/admin/editar_motorista.html', {
        'form': form,
        'motorista': motorista
    })

@login_required
def listar_veiculos(request):
    """Lista todos os veículos"""
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Acesso negado!')
        return redirect('motorista_dashboard')
    
    veiculos = Veiculo.objects.order_by('placa')
    
    return render(request, 'ponto/admin/veiculos.html', {
        'veiculos': veiculos
    })

@login_required
def cadastrar_veiculo(request):
    """Cadastra novo veículo"""
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Acesso negado!')
        return redirect('motorista_dashboard')
    
    if request.method == 'POST':
        form = VeiculoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Veículo cadastrado com sucesso!')
            return redirect('listar_veiculos')
    else:
        form = VeiculoForm()
    
    return render(request, 'ponto/admin/cadastrar_veiculo.html', {
        'form': form
    })

@login_required
def editar_veiculo(request, id):
    """Edita veículo existente"""
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Acesso negado!')
        return redirect('motorista_dashboard')
    
    veiculo = get_object_or_404(Veiculo, id=id)
    
    if request.method == 'POST':
        form = VeiculoForm(request.POST, instance=veiculo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Veículo atualizado com sucesso!')
            return redirect('listar_veiculos')
    else:
        form = VeiculoForm(instance=veiculo)
    
    return render(request, 'ponto/admin/editar_veiculo.html', {
        'form': form,
        'veiculo': veiculo
    })

@login_required
def listar_mercados(request):
    """Lista todos os mercados"""
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Acesso negado!')
        return redirect('motorista_dashboard')
    
    mercados = Mercado.objects.order_by('nome')
    
    return render(request, 'ponto/admin/mercados.html', {
        'mercados': mercados
    })

@login_required
def cadastrar_mercado(request):
    """Cadastra novo mercado"""
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Acesso negado!')
        return redirect('motorista_dashboard')
    
    if request.method == 'POST':
        form = MercadoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Mercado cadastrado com sucesso!')
            return redirect('listar_mercados')
    else:
        form = MercadoForm()
    
    return render(request, 'ponto/admin/cadastrar_mercado.html', {
        'form': form
    })

@login_required
def editar_mercado(request, id):
    """Edita mercado existente"""
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Acesso negado!')
        return redirect('motorista_dashboard')
    
    mercado = get_object_or_404(Mercado, id=id)
    
    if request.method == 'POST':
        form = MercadoForm(request.POST, instance=mercado)
        if form.is_valid():
            form.save()
            messages.success(request, 'Mercado atualizado com sucesso!')
            return redirect('listar_mercados')
    else:
        form = MercadoForm(instance=mercado)
    
    return render(request, 'ponto/admin/editar_mercado.html', {
        'form': form,
        'mercado': mercado
    })

@login_required
def listar_registros(request):
    """Lista todos os registros com filtros"""
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Acesso negado!')
        return redirect('motorista_dashboard')
    
    registros = RegistroPonto.objects.select_related(
        'motorista', 'motorista__veiculo', 'motorista__mercado'
    ).order_by('-data_hora')
    
    # Filtros opcionais
    motorista_id = request.GET.get('motorista')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    tipo = request.GET.get('tipo')
    
    if motorista_id:
        registros = registros.filter(motorista_id=motorista_id)
    
    if data_inicio:
        registros = registros.filter(data_hora__date__gte=data_inicio)
    
    if data_fim:
        registros = registros.filter(data_hora__date__lte=data_fim)
    
    if tipo:
        registros = registros.filter(tipo=tipo)
    
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(registros, 25)  # 25 registros por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    motoristas = Motorista.objects.filter(ativo=True).order_by('nome_completo')
    
    return render(request, 'ponto/admin/registros.html', {
        'page_obj': page_obj,
        'motoristas': motoristas,
        'filtros': {
            'motorista_id': motorista_id,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'tipo': tipo,
        }
    })

@login_required
def detalhe_registro(request, id):
    """API endpoint para detalhes do registro"""
    if not (request.user.is_superuser or request.user.is_staff):
        return JsonResponse({'error': 'Acesso negado'}, status=403)
    
    try:
        registro = RegistroPonto.objects.select_related(
            'motorista', 'motorista__veiculo'
        ).get(id=id)
        
        data = {
            'success': True,
            'registro': {
                'motorista': registro.motorista.nome_completo,
                'tipo': registro.get_tipo_display(),
                'data_hora': registro.data_hora.strftime('%d/%m/%Y %H:%M'),
                'veiculo': str(registro.motorista.veiculo),
                'km_odometro': registro.km_odometro,
                'nivel_combustivel': registro.nivel_combustivel,
                'observacoes': registro.observacoes or '',
                'km_rodados': registro.calcular_km_rodados() if registro.tipo == 'saida' else 0,
                'horas_trabalhadas': round(registro.calcular_horas_trabalhadas(), 2) if registro.tipo == 'saida' else 0
            }
        }
        
        return JsonResponse(data)
        
    except RegistroPonto.DoesNotExist:
        return JsonResponse({'error': 'Registro não encontrado'}, status=404)
    
@login_required
def detalhe_registro_html(request, id):
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Acesso negado!')
        return redirect('motorista_dashboard')

    registro = get_object_or_404(RegistroPonto, id=id)

    return render(request, 'ponto/admin/detalhe_registro.html', {
        'registro': registro
    })


@login_required
def exportar_relatorio(request):
    """Exporta relatório em Excel"""
    if not (request.user.is_superuser or request.user.is_staff):
        return JsonResponse({'error': 'Acesso negado'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    
    try:
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill, Alignment
        
        data = json.loads(request.body)
        
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório de Ponto"
        
        # Cabeçalhos
        headers = [
            'Data', 'Motorista', 'CPF', 'Veículo', 'Mercado',
            'Entrada', 'Saída', 'Horas Trabalhadas', 'KM Rodados', 'Valor Dia'
        ]
        
        # Estilo do cabeçalho
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Dados (implementar busca de dados conforme gerar_relatorio)
        # ...
        
        # Preparar response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=relatorio_ponto_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def api_status_motoristas_hoje(request):
    """API endpoint para status dos motoristas hoje"""
    if not (request.user.is_superuser or request.user.is_staff):
        return JsonResponse({'error': 'Acesso negado'}, status=403)
    
    hoje = timezone.now().date()
    motoristas = Motorista.objects.filter(ativo=True).select_related('veiculo', 'mercado')
    
    dados_motoristas = []
    
    for motorista in motoristas:
        entrada_hoje = RegistroPonto.objects.filter(
            motorista=motorista,
            tipo='entrada',
            data_hora__date=hoje
        ).first()
        
        saida_hoje = RegistroPonto.objects.filter(
            motorista=motorista,
            tipo='saida',
            data_hora__date=hoje
        ).first()
        
        if entrada_hoje and saida_hoje:
            status = 'finalizado'
        elif entrada_hoje:
            status = 'trabalhando'
        else:
            status = 'nao_iniciou'
        
        dados_motoristas.append({
            'id': motorista.id,
            'nome': motorista.nome_completo,
            'veiculo': str(motorista.veiculo),
            'mercado': motorista.mercado.nome,
            'status': status,
            'entrada_hoje': entrada_hoje.hora_formatada if entrada_hoje else None,
            'saida_hoje': saida_hoje.hora_formatada if saida_hoje else None,
        })
    
    return JsonResponse({
        'success': True,
        'motoristas': dados_motoristas
    })

@login_required
def api_registro_fotos(request, id):
    """API endpoint para fotos do registro"""
    if not (request.user.is_superuser or request.user.is_staff):
        return JsonResponse({'error': 'Acesso negado'}, status=403)
    
    try:
        registro = RegistroPonto.objects.get(id=id)
        
        data = {
            'success': True,
            'fotos': {
                'odometro': registro.foto_odometro.url if registro.foto_odometro else None,
                'combustivel': registro.foto_combustivel.url if registro.foto_combustivel else None,
            }
        }
        
        return JsonResponse(data)
        
    except RegistroPonto.DoesNotExist:
        return JsonResponse({'error': 'Registro não encontrado'}, status=404)

@login_required
def relatorio_ponto(request):
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Acesso negado!')
        return redirect('motorista_dashboard')

    motoristas = Motorista.objects.filter(ativo=True).order_by('nome_completo')
    veiculos = Veiculo.objects.order_by('placa')

    registros = []
    filtro_aplicado = False

    if request.method == 'GET':
        data_inicio = request.GET.get('data_inicio')
        data_fim = request.GET.get('data_fim')
        motorista_id = request.GET.get('motorista', '')
        veiculo_id = request.GET.get('veiculo', '')

        # Converte para string para facilitar comparação no template
        motorista_id = str(motorista_id)
        veiculo_id = str(veiculo_id)

        if data_inicio and data_fim:
            filtro_aplicado = True
            registros_query = RegistroPonto.objects.filter(
                data_hora__date__gte=data_inicio,
                data_hora__date__lte=data_fim,
            )

            if motorista_id:
                registros_query = registros_query.filter(motorista_id=motorista_id)

            if veiculo_id:
                registros_query = registros_query.filter(motorista__veiculo_id=veiculo_id)

            registros_query = registros_query.select_related(
                'motorista', 'motorista__veiculo', 'motorista__mercado'
            ).order_by('motorista', 'data_hora')

            temp = defaultdict(
                lambda: {'entrada': None, 'saida': None, 'registro_entrada': None, 'registro_saida': None}
            )

            for reg in registros_query:
                chave = (reg.motorista.id, reg.data_hora.date())
                if reg.tipo == 'entrada':
                    temp[chave]['entrada'] = reg.data_hora
                    temp[chave]['registro_entrada'] = reg
                else:
                    temp[chave]['saida'] = reg.data_hora
                    temp[chave]['registro_saida'] = reg

            registros = temp.values()

    context = {
        'motoristas': motoristas,
        'veiculos': veiculos,
        'registros': registros,
        'filtro_aplicado': filtro_aplicado,
        'data_inicio': request.GET.get('data_inicio', ''),
        'data_fim': request.GET.get('data_fim', ''),
        'motorista_id': motorista_id,
        'veiculo_id': veiculo_id,
    }

    return render(request, 'ponto/admin/relatorio_ponto.html', context)

@login_required
def gerar_relatorio(request):
    """Gera relatório personalizado (API para AJAX ou POST)"""
    if not (request.user.is_superuser or request.user.is_staff):
        return JsonResponse({'error': 'Acesso negado'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    
    try:
        data = json.loads(request.body)
        # Filtros
        data_inicio = data.get('data_inicio')
        data_fim = data.get('data_fim')
        motorista_id = data.get('motorista')
        veiculo_id = data.get('veiculo')
        
        registros = RegistroPonto.objects.select_related(
            'motorista', 'motorista__veiculo'
        ).filter(
            data_hora__date__gte=data_inicio,
            data_hora__date__lte=data_fim
        )
        if motorista_id:
            registros = registros.filter(motorista_id=motorista_id)
        if veiculo_id:
            registros = registros.filter(motorista__veiculo_id=veiculo_id)
        
        relatorio_data = []
        for registro in registros:
            relatorio_data.append({
                'data': registro.data_hora.strftime('%d/%m/%Y'),
                'motorista': registro.motorista.nome_completo,
                'cpf': registro.motorista.cpf,
                'veiculo': str(registro.motorista.veiculo),
                'mercado': registro.motorista.mercado.nome if registro.motorista.mercado else '',
                'entrada': registro.hora_formatada if registro.tipo == 'entrada' else '',
                'saida': registro.hora_formatada if registro.tipo == 'saida' else '',
                'horas_trabalhadas': registro.calcular_horas_trabalhadas() if registro.tipo == 'saida' else '',
                'km_rodados': registro.calcular_km_rodados() if registro.tipo == 'saida' else '',
                'valor_dia': float(registro.motorista.valor_dia),
            })
        
        return JsonResponse({'success': True, 'data': relatorio_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
@login_required
def exportar_relatorio_excel(request):
    if not (request.user.is_superuser or request.user.is_staff):
        return HttpResponse('Acesso negado', status=403)

    # Receber filtros da requisição GET
    data_inicio = request.GET.get('data_inicio', '').strip()
    data_fim = request.GET.get('data_fim', '').strip()
    motorista_id = request.GET.get('motorista')
    veiculo_id = request.GET.get('veiculo')

    registros_query = RegistroPonto.objects.all()

    # Aplicar filtros de data somente se válidos
    if data_inicio:
        registros_query = registros_query.filter(data_hora__date__gte=data_inicio)
    if data_fim:
        registros_query = registros_query.filter(data_hora__date__lte=data_fim)

    if motorista_id:
        registros_query = registros_query.filter(motorista_id=motorista_id)
    if veiculo_id:
        registros_query = registros_query.filter(motorista__veiculo_id=veiculo_id)

    registros_query = registros_query.select_related('motorista', 'motorista__veiculo', 'motorista__mercado').order_by('motorista', 'data_hora')

    temp = defaultdict(lambda: {'entrada': None, 'saida': None, 'registro_entrada': None, 'registro_saida': None})

    for reg in registros_query:
        chave = (reg.motorista.id, reg.data_hora.date())
        if reg.tipo == 'entrada':
            temp[chave]['entrada'] = reg.data_hora
            temp[chave]['registro_entrada'] = reg
        else:
            temp[chave]['saida'] = reg.data_hora
            temp[chave]['registro_saida'] = reg

    registros = temp.values()

    # Criar Workbook do Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Ponto"

    # Cabeçalhos
    headers = [
        'Motorista', 'CPF', 'Veículo', 'Mercado', 'Data', 
        'Entrada', 'Saída', 'Horas Trabalhadas', 'KM Rodados', 'Valor Dia'
    ]

    # Estilo do cabeçalho
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    for reg in registros:
        motor = reg['registro_entrada'].motorista if reg['registro_entrada'] else (reg['registro_saida'].motorista if reg['registro_saida'] else None)
        data = reg['entrada'].date() if reg['entrada'] else (reg['saida'].date() if reg['saida'] else '')
        entrada_hora = reg['entrada'].strftime('%H:%M') if reg['entrada'] else ''
        saida_hora = reg['saida'].strftime('%H:%M') if reg['saida'] else ''
        horas_trabalhadas = ''
        km_rodados = ''
        valor_dia = ''

        if reg['registro_entrada'] and reg['registro_saida']:
            horas_trabalhadas = round(reg['registro_saida'].calcular_horas_trabalhadas(), 2)
            km_rodados = round(reg['registro_saida'].calcular_km_rodados(), 2)
            valor_dia = float(motor.valor_dia) if motor and motor.valor_dia else ''

        ws.cell(row=row_num, column=1, value=motor.nome_completo if motor else '')
        ws.cell(row=row_num, column=2, value=motor.cpf if motor else '')
        ws.cell(row=row_num, column=3, value=str(motor.veiculo) if motor and motor.veiculo else '')
        ws.cell(row=row_num, column=4, value=motor.mercado.nome if motor and motor.mercado else '')
        ws.cell(row=row_num, column=5, value=data)
        ws.cell(row=row_num, column=6, value=entrada_hora)
        ws.cell(row=row_num, column=7, value=saida_hora)
        ws.cell(row=row_num, column=8, value=horas_trabalhadas)
        ws.cell(row=row_num, column=9, value=km_rodados)
        ws.cell(row=row_num, column=10, value=valor_dia)

        row_num += 1

    # Ajustar largura das colunas automaticamente (simplificado)
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"relatorio_ponto_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response